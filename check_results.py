"""
03. 03. 2025
Author - Bhagya Wijeratne

Input - File containing the string Id's to be matched
      - Excel file of the hub proteins related to the trait
Output - Updated excel file where the genes from the provided list are highlighted in green

Identifying the positive Kranz Regulators present in the obtained hub proteins
#################################################################################################
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def load_id_list(id_file):
    with open(id_file, "r") as f:
        return {line.strip().upper() for line in f if line.strip()}  # Read and clean IDs


# CHANGE: sheet names when coloring the internal hubs
def highlight_matching_ids(excel_file, sheet_name, id_list, output_file):
    # Load the workbook
    wb = load_workbook(excel_file)
    # green = 90EE90 yellow = FFFF00
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    modified_cells = 0

    # for sheet_name in sheet_names:  # CHANGE: uncomment for internal hubs
    # Load the sheet data
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    ws = wb[sheet_name]

    # # CHANGE: uncomment for internal hubs
    # # Check if "String id" column exists
    # if "String id" not in df.columns:
    #     print(f"Error: 'String id' column not found in sheet {sheet_name}.")
    #     continue

    # # CHANGE: uncomment for internal hubs
    # # Iterate through rows and highlight matching IDs
    # for index, row in df.iterrows():
    #     if row["String id"].upper() in id_list:
    #         modified_cells += 1
    #         for cell in ws[index + 2]:  # +2 to account for header (Excel rows are 1-based)
    #             cell.fill = highlight_fill

    # CHANGE: uncomment for clusters
    # Iterate through all cells and highlight matching ones
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
                                  start=2):
        for cell in row:
            if str(cell.value) in id_list:
                cell.fill = highlight_fill
                modified_cells += 1

        # Save the updated file
        wb.save(output_file)
        # CHANGE: uncomment below accordingly
        # print(f"Highlighted rows saved in {output_file}.")
        print(f"Highlighted matching cells in {output_file}")
        print(f"Total modified cells: {modified_cells}")


excel_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\results\hub_analysis\yellow\hub_analysis_yellow_1_edited_I.xlsx"  # Change this to your actual file
# CHANGE: fix this accordingly
sheet_names = "clustered proteins"  # CHANGE: List of sheet names ["inter-hubs", "intra-hubs"]
output_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\results\hub_analysis\yellow\hub_analysis_yellow_1_edited_I.xlsx"
# CHANGE: pick the correct file
# id_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\paper results\positive_regulators_with_prefix.txt"  # Text file containing IDs
# id_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\paper results\positive_regulators_without_prefix.txt"  # Text file containing IDs
id_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\paper results\negative_regulators_without_prefix.txt"  # Text file containing IDs
# id_file = r"C:\Users\BhagyaWijeratne\Desktop\Level IV\Research\methodolgy\PPI network\paper results\negative regulators_with_prefix.txt"  # Text file containing IDs

# calling functions
id_list = load_id_list(id_file)  # Load IDs from file with prefix
# CHANGE: change to sheet names for the internal hubs
highlight_matching_ids(excel_file, sheet_names, id_list, output_file)
