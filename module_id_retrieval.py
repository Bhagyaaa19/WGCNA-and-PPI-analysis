# 25/12/2024
# Author - Bhagya Wijeratne

# Input: The txt input files to be converted into an excel file
# Output: Writing to the excel file to convert the tsv file to an excel for easy retrieval

# Creates an excel file of all the genes for all the input txt files

# use cases
# used to convert NCBI output for module genes mapping from maize ID to gene ID into an excel file
# ==========================================================================================

import pandas as pd
from openpyxl import load_workbook

# file paths
# "C:\\Users\\BhagyaWijeratne\\Desktop\\Level IV\\Research\\methodolgy\\CoExpNetwork\\Green_sample_classification_gene_ids\\green_sample_classification.txt"
txt_file = "C:\\Users\\BhagyaWijeratne\\Desktop\\Level IV\\Research\\methodolgy\\CoExpNetwork\\blue_gene_ids\\blue_gene_ids_batch_entrez.txt"
# "C:\\Users\\BhagyaWijeratne\\Desktop\\Level IV\\Research\\methodolgy\\CoExpNetwork\\Green_sample_classification_gene_ids\\green_NCBI_output_genes_part2.txt"
# "C:\\Users\\BhagyaWijeratne\\Desktop\\Level IV\\Research\\methodolgy\\CoExpNetwork\\Turquoise_sample_classficiation_gene_ids\\turquoise_NCBI_output_genes.txt"
excel_file = "C:\\Users\\BhagyaWijeratne\\Desktop\\Level IV\\Research\\methodolgy\\CoExpNetwork\\geneInfo_sample_classification.xlsx"
sheet_name = "Mapped_blue_module_genes"


def write_text_to_excel(txt_file, excel_file, sheet_name):
    # Read the tab-separated data from the text file
    df = pd.read_csv(txt_file, sep='\t')  # Automatically splits columns by tabs

    # Check if the Excel file exists
    try:
        # Load the existing workbook
        book = load_workbook(excel_file)
        # Use ExcelWriter with mode='a' to append data
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Write the DataFrame to the specified sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False,
                        startrow=writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0)
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


print("Data successfully written to the excel")
write_text_to_excel(txt_file, excel_file, sheet_name)
